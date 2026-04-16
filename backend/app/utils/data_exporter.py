"""
Exportador de datos en múltiples formatos
Soporte para CSV, Excel, PDF y JSON
"""
import csv
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
from io import StringIO, BytesIO


def export_to_csv(data: List[Dict], columns: Optional[List[str]] = None) -> str:
    if not data:
        return ""
    
    output = StringIO()
    
    # Determinar columnas
    if columns is None:
        columns = list(data[0].keys())
    
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    
    for row in data:
        # Convertir valores complejos a string
        row_clean = {}
        for key in columns:
            value = row.get(key, '')
            if isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, (dict, list)):
                value = json.dumps(value)
            row_clean[key] = value
        
        writer.writerow(row_clean)
    
    return output.getvalue()


def export_to_json(data: List[Dict], indent: int = 2, include_metadata: bool = True) -> str:
    if include_metadata:
        output = {
            'metadata': {
                'exported_at': datetime.now().isoformat(),
                'record_count': len(data),
                'version': '1.0'
            },
            'data': data
        }
    else:
        output = data
    
    return json.dumps(output, indent=indent, default=str)


def export_to_excel_binary(data: List[Dict], sheet_name: str = 'Datos') -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
    
    if not data:
        return b""
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Escribir headers
    columns = list(data[0].keys())
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Escribir datos
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            if isinstance(value, datetime):
                ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(value))
    
    # Ajustar ancho de columnas
    for col_idx, col_name in enumerate(columns, 1):
        max_length = max(
            len(str(col_name)),
            max((len(str(row.get(col_name, ''))) for row in data), default=0)
        )
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)
    
    # Guardar en buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_report_pdf(
    title: str,
    sections: List[Dict],
    metadata: Optional[Dict] = None
) -> bytes:
    # Placeholder - implementación real requiere librerías adicionales
    # Ejemplo con reportlab:
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    
    # Título
    c.setFont("Helvetica-Bold", 20)
    c.drawString(50, 750, title)
    
    # Secciones
    y_position = 700
    for section in sections:
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y_position, section['title'])
        y_position -= 20
        
        c.setFont("Helvetica", 11)
        for line in section['content']:
            c.drawString(50, y_position, str(line))
            y_position -= 15
        
        y_position -= 20
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue()
    """
    
    # Versión simplificada sin dependencias
    content = f"REPORTE: {title}\n"
    content += "=" * 50 + "\n\n"
    
    for section in sections:
        content += f"{section['title']}\n"
        content += "-" * 30 + "\n"
        for item in section.get('content', []):
            content += f"  • {item}\n"
        content += "\n"
    
    if metadata:
        content += "\nMetadatos:\n"
        for key, value in metadata.items():
            content += f"  {key}: {value}\n"
    
    return content.encode('utf-8')


def export_financial_summary(
    period: str,
    summary: Dict,
    format: str = 'json'
) -> str:
    # Preparar datos
    data = {
        'period': period,
        'generated_at': datetime.now().isoformat(),
        'summary': summary
    }
    
    if format == 'json':
        return json.dumps(data, indent=2, default=str)
    
    elif format == 'csv':
        # Aplanar datos para CSV
        rows = []
        for category, value in summary.items():
            rows.append({
                'period': period,
                'category': category,
                'value': value,
                'generated_at': data['generated_at']
            })
        
        return export_to_csv(rows, columns=['period', 'category', 'value', 'generated_at'])
    
    else:
        raise ValueError(f"Formato no soportado: {format}")


def create_data_package(
    user_id: int,
    data_types: List[str],
    data: Dict[str, List[Dict]]
) -> Dict:
    package = {
        'user_id': user_id,
        'export_timestamp': datetime.now().isoformat(),
        'data_types_included': data_types,
        'data': {},
        'summary': {
            'total_records': 0,
            'by_type': {}
        }
    }
    
    total_records = 0
    for data_type in data_types:
        if data_type in data:
            records = data[data_type]
            package['data'][data_type] = records
            record_count = len(records)
            package['summary']['by_type'][data_type] = record_count
            total_records += record_count
    
    package['summary']['total_records'] = total_records
    
    return package


def compress_data_package(package: Dict) -> bytes:
    import zipfile
    
    buffer = BytesIO()
    
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Agregar manifiesto
        manifest = {
            'created_at': datetime.now().isoformat(),
            'contents': list(package.get('data', {}).keys()),
            'record_count': package.get('summary', {}).get('total_records', 0)
        }
        zip_file.writestr(
            'manifest.json',
            json.dumps(manifest, indent=2, default=str)
        )
        
        # Agregar datos por tipo
        for data_type, records in package.get('data', {}).items():
            # JSON version
            zip_file.writestr(
                f'{data_type}.json',
                json.dumps(records, indent=2, default=str)
            )
            
            # CSV version
            if records:
                csv_content = export_to_csv(records)
                zip_file.writestr(f'{data_type}.csv', csv_content)
    
    buffer.seek(0)
    return buffer.getvalue()
